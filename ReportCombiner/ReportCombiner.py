## https://steam.oxxostudio.tw/category/python/example/excel-read.html 
## https://ithelp.ithome.com.tw/articles/10246377
## https://hackmd.io/@amostsai/SJkC1_EcX?type=view
## https://www.jb51.net/article/237279.htm
import os
import math
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border,Side
import csv
#from openpyxl.chart import LineChart, Reference
##import matplotlib.pyplot as plt
#from math import ceil
##===========================#    
#def fill_data_into_excel(r, c, data,sheet):
#    sheet.cell(r, c).value = data
#  
#def Init_Excel_Table(row, sheet):
#    colInfo = {}
#    for i in range(0,len(row)):
#        if(row[i] == "start_time"):
#            colInfo["start_time"] = (i)
#            sheet.cell(1, COL_INDEX_TIME).value = 'start_time'
#        elif(row[i] == "channel"):
#            colInfo["channel"] = (i)
#        #    sheet.write(0, COL_INDEX_CHANL_L, 'channel')
#        elif(row[i] == "data"):
#            colInfo["data"] = (i) 
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_HEX_L, 'HEX_24_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_HEX_R, 'HEX_24_R', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_DEC_L, 'DEC_24_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_DEC_R, 'DEC_24_R', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_16bit_DEC_L, 'DEC_16_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_16bit_DEC_R, 'DEC_16_R', sheet) 
#            #sheet.cell(1, COL_INDEX_DATA_24bit_HEX_L).value = 'HEX_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_HEX_R).value = 'HEX_24_R'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_DEC_L).value = 'DEC_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_DEC_R).value = 'DEC_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_16bit_DEC_L).value = 'DEC_16_L'
#            #sheet.cell(1, COL_INDEX_DATA_16bit_DEC_R).value = 'DEC_16_L'    
#    
#    fill_data_into_excel(1, COL_INDEX_SAMPLING_RATE, 'SamplingRate(Hz)', sheet)
#    #fill_data_into_excel(4, COL_INDEX_WAVEFORM, 'Waveform(L)', sheet)
#    
#    return colInfo;
#
#def convert_hexstr2hexval(valStr):
#    CONST_SIGN_BIT      = 0x800000
#    CONST_SIGN_VAL_MAX  = 0x7FFFFF
#    
#    try:
#        val_24 = int(valStr, 16)
#    except:
#        print(valStr)
#    if(val_24 & CONST_SIGN_BIT):
#        val_24 = (val_24 & CONST_SIGN_VAL_MAX) - CONST_SIGN_VAL_MAX - 1 
#    return val_24
#    
#
#def raw_convert_bytearr(rawList):
#    retList = []
#    for i in range(0, len(rawList)):
#        retList.append( rawList[i] & 0xFF)
#        retList.append( (rawList[i] >> 8) & 0xFF ) 
#    #print(retList)
#    return bytearray(retList)
#    
#def dual_raw_convert_bytearr(rawList1, rawList2):
#    retList = []
#    for i in range(0, len(rawList1)):
#        retList.append( rawList1[i] & 0xFF)
#        retList.append( (rawList1[i] >> 8) & 0xFF ) 
#        retList.append( rawList2[i] & 0xFF)
#        retList.append( (rawList2[i] >> 8) & 0xFF )        
#    #print(retList)
#    return bytearray(retList)
#
#def data_plot(ws, dataLen):
#    chart = LineChart()
#    chart.title = "Data - left, 24bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_24bit_DEC_L, min_row=1, max_col=COL_INDEX_DATA_24bit_DEC_L, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)
#    ws.add_chart(chart, "I4")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - right, 24bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_24bit_DEC_R, min_row=1, max_col=COL_INDEX_DATA_24bit_DEC_R, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)    
#    ws.add_chart(chart, "I20")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - left, 16bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_16bit_DEC_L, min_row=1, max_col=COL_INDEX_DATA_16bit_DEC_L, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)
#    ws.add_chart(chart, "I36")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - right, 16bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_16bit_DEC_R, min_row=1, max_col=COL_INDEX_DATA_16bit_DEC_R, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)    
#    ws.add_chart(chart, "I52")   
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10    
#

def csv_to_excel(csv_filename, excel_filename):

    # Read CSV file
    csv_data = []
    with open(csv_filename) as f:
        csv_data = [row for row in csv.reader(f)]
    
    # Write to Excel file
    workbook = openpyxl.workbook.Workbook()
    ws = workbook.active
    
    lineCount = 1 
    numCol = 0
    sepSymbol = ""
    writeRow = 1
    keepReading = False
    readStr = ""
    for row in csv_data:
        if(lineCount == 1):
            sepSymbol = row[0][4]   #Assume the string begins with "sep=" 
        elif(lineCount == 2):
            numCol = row[0].count(';') +1
            #print(row[0],"\n")
        else:
            if(keepReading == False):
                readStr = ""
            
            for i in range(0,len(row)):
                readStr = readStr + row[i]
            readStr = readStr.replace("; ",". ")

            #if(writeRow >345 and writeRow <365):
            #    print(writeRow, keepReading, readStr.count(';'))
            #    print(readStr, "\n")
                
            if((keepReading == False) and (readStr.count(';') == 0)):
                keepReading = False
            elif((keepReading == False) and (readStr.count(';') +1 < numCol) ):
                keepReading = True
            elif((keepReading == True) and (readStr.count(';') +1 < numCol) ):
                keepReading = True
            else:
                keepReading = False
                readStr = readStr.split(";")
                #print(readStr)
                for c in range(0,len(readStr)):
                    ws[chr(65+c)+str(writeRow)].value =  readStr[c]
                writeRow = writeRow+1
            #print(lineCount, readStr)

            
        lineCount = lineCount+1
       
        #print(row)
        
        
        
    workbook.save(excel_filename)        
        
def find_info_from_report(itemName,wb_harmony, wb_ebq):
    ws_harmony = wb_harmony["TestPlan"]
    readRowIdx = 1
    while(True):
        readName = ws_harmony.cell(readRowIdx,1).value
        if(readName == itemName):
            if(ws_harmony.cell(readRowIdx,10).value == "Pass"):
                return ws_harmony.cell(readRowIdx,10).value
            else:
                break;
        
        readRowIdx += 1
        if(readRowIdx > 5000):
            break

    ws_ebq = wb_ebq["Sheet"]
    readRowIdx = 1
    while(True):
        readName = ws_ebq.cell(readRowIdx,1).value
        if(itemName in readName):
            if(ws_ebq.cell(readRowIdx,5).value == "Passed"):
                return "Pass_EBQ"
            else:
                return ws_ebq.cell(readRowIdx,5).value
        
        readRowIdx += 1
        if(readRowIdx > 5000):
            break
            
#    for row in csvInRowEbq:
#        print(row)
#        print("\n\n")
    
#    while(True):
#        readName = ws_harmony.cell(readRowIdx,1).value
#        if(readName == itemName):
#            if(ws_harmony.cell(readRowIdx,10).value == "Pass"):
#                return ws_harmony.cell(readRowIdx,10).value
#            else:
#                break;
#        
#        readRowIdx += 1
#        if(readRowIdx > 5000):
#            break    
    
        
def AppStart():
    inConfigF = open("sourceFile.config", 'r')
    inConfigList = inConfigF.readlines()
    
#-----------------------------------------------------#    
    originalFile = inConfigList[0].split()[1]
    harmonyReport = inConfigList[1].split()[1]
    ebqReport = inConfigList[2].split()[1]
    
    sheetInfo = [[inConfigList[3].split()[0], 0],
                 [inConfigList[4].split()[0], 0],
                 [inConfigList[5].split()[0], 0],
                 [inConfigList[6].split()[0], 0],
                 [inConfigList[7].split()[0], 0]
                 ]
    
    wb_original = openpyxl.load_workbook(originalFile)
    wb_harmony = openpyxl.load_workbook(harmonyReport)
    
    # xlsx file,converted from CSV
    csvfile_ebq = open(ebqReport)
    csv_to_excel(ebqReport, "Tmp.xlsx")
    wb_ebq = openpyxl.load_workbook("Tmp.xlsx")
    #print(inConfigList[0].split())
    #print(originalFile)
    #print(harmonyReport)
    #print(ebqReport)
    #print(sheetInfo)
    
#    csvfile_ebq = open(ebqReport)
#    csvInRowEbq = csv.reader(csvfile_ebq)     # 讀取 csv 檔案
#    
#    for row in csvInRowEbq:
#        row1 = row[0]
#        #row = row[0].split(sep=";")
#        print(len(row))
#        print((row))
#        print("\n\n")
#    exit()

    
#-----------------------------------------------------# 
# init the final report
    finalReportName = "FinalReport.xlsx"
    
    workbookOutput = openpyxl.Workbook()
    ss_sheet = workbookOutput['Sheet']
    ss_sheet.title = 'Summary'
    workbookOutput.create_sheet("1_"+sheetInfo[0][0], 2)
    workbookOutput.create_sheet("2_"+sheetInfo[1][0], 3)
    workbookOutput.create_sheet("3_"+sheetInfo[2][0], 4)
    workbookOutput.create_sheet("4_"+sheetInfo[3][0], 5)
    workbookOutput.create_sheet("5_"+sheetInfo[4][0], 6)
    
    workbookOutput.active = 0
    ws = workbookOutput.active
    ws['B2'].value =  "Report Type"
    ws['B3'].value =  "Testing Purpose"
    ws['B4'].value =  "Release version"
    ws['B5'].value =  "Interface"
    ws['B6'].value =  "Host Platform"
    ws['B7'].value =  "Host OS"
    ws['B8'].value =  "Report location"
    
    ws['B12'].value =  "Suites Breakdown"
    ws['C12'].value =  "Total"
    ws['D12'].value =  "Pass"
    ws['E12'].value =  "Fail"
    ws['F12'].value =  "Inconclusive"
    ws['G12'].value =  "NA"
    ws['H12'].value =  ""
    ws['I12'].value =  "Completed[%]"
    ws['J12'].value =  "Pass[%]"
    ws['K12'].value =  "Bug Count"

    ws['B13'].value = sheetInfo[0][0]
    ws['B14'].value = sheetInfo[1][0]
    ws['B15'].value = sheetInfo[2][0] 
    ws['B16'].value = sheetInfo[3][0]
    ws['B17'].value = sheetInfo[4][0]     
    
    ws['B20'].value = "JIRA" 
    ws['C20'].value = "Summary"						
    ws['I20'].value = "Priority"
    ws['J20'].value = "Status"
    ws['K20'].value = "Assignee"
    
    ws.column_dimensions['B'].width = 16.5
    ws.column_dimensions['C'].width = 10.0
    ws.column_dimensions['D'].width = 10.0
    ws.column_dimensions['E'].width = 10.0
    ws.column_dimensions['F'].width = 10.0
    ws.column_dimensions['G'].width = 8.0
    ws.column_dimensions['I'].width = 14.0
    ws.column_dimensions['J'].width = 14.0
    ws.column_dimensions['K'].width = 14.0
    ws.row_dimensions[4].height = 40.0
    
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=11)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=11)
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=11)
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=11)
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=11)
    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=11)
    ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=11)
    ws.merge_cells(start_row=20, start_column=3, end_row=20, end_column=8)
    ws.merge_cells(start_row=21, start_column=3, end_row=21, end_column=8)
    ws.merge_cells(start_row=22, start_column=3, end_row=22, end_column=8)
    ws.merge_cells(start_row=23, start_column=3, end_row=23, end_column=8)
    ws.merge_cells(start_row=24, start_column=3, end_row=24, end_column=8)
    ws.merge_cells(start_row=25, start_column=3, end_row=25, end_column=8)

    border = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000'))

    for c in range(2,12):                    
        for r in range(2,9):
            ws[chr(65+c-1)+str(r)].border = border
        for r in range(12,18):
            ws[chr(65+c-1)+str(r)].border = border
        for r in range(20,26):
            ws[chr(65+c-1)+str(r)].border = border
            
        

    for i in range(1,6):
        workbookOutput.active = i
        ws = workbookOutput.active    
        ws['A1'].value =  "Item"
        ws['B1'].value =  "Description"
        ws['C1'].value =  "TestPlatform"
        ws['D1'].value =  "Result"
        ws['E1'].value =  "JIRA"        #Filled by hand if needed, QA 
        ws['F1'].value =  "Reason"      #Filled by hand if needed, FW
        ws['G1'].value =  "Pass rate"   #Filled by hand
        ws['H1'].value =  "New Item"    #Filled by hand if needed, QA
        ws.column_dimensions['A'].width = 25.0
        ws.column_dimensions['B'].width = 30.0
        ws.column_dimensions['C'].width = 12.0
        ws.column_dimensions['D'].width = 12.0
        ws.column_dimensions['E'].width = 12.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 8.0
        ws.column_dimensions['H'].width = 9.0
#-----------------------------------------------------#
    ws_original = wb_original["TestPlan"]
#-----------------------------------------------------#
    readRowIdx = 1
    for i in range(0,5):
        while(True):
            if(ws_original.cell(readRowIdx,1).value == None):
                readRowIdx = readRowIdx+1   
            elif(sheetInfo[i][0] in ws_original.cell(readRowIdx,1).value):
                sheetInfo[i][1] = readRowIdx + 2
                readRowIdx = readRowIdx+2
                break
            else:
                readRowIdx = readRowIdx+1
    
#-----------------------------------------------------#  
    fontFail = Font(color='9c0006')
    fontInc = Font(color='006100')
    for i in range(0,5):
        statResult = [0,0,0,0,0]    #Total / Pass / Fail / Inconclusive / NA
        readRowIdx = sheetInfo[i][1]
        
        workbookOutput.active = i + 1
        ws = workbookOutput.active 
        wrRowIdx = 2

        while(True):
            #print(ws_original.cell(readRowIdx,1).value)
            #Item
            ws['A'+str(wrRowIdx)] = ws_original.cell(readRowIdx,1).value
            #Description
            ws['B'+str(wrRowIdx)] = ws_original.cell(readRowIdx,2).value
            #TestPlatform
            ws['C'+str(wrRowIdx)] = ws_original.cell(readRowIdx,5).value
            #Result, TBD
            ws['D'+str(wrRowIdx)] = find_info_from_report(ws_original.cell(readRowIdx,1).value, wb_harmony, wb_ebq)
            if("Fail" in ws['D'+str(wrRowIdx)].value):
                ws['D'+str(wrRowIdx)].font = fontFail
            elif("Inconclusive" in ws['D'+str(wrRowIdx)].value):
                ws['D'+str(wrRowIdx)].font = fontInc
                
            readRowIdx += 1
            wrRowIdx += 1

            if(ws_original.cell(readRowIdx,1).value == None):
                #print(readRowIdx, ws_original.cell(readRowIdx,1).value)
                break;
                

        for c in range(1,9):
            for r in range(1,wrRowIdx):
                ws[chr(65+c-1)+str(r)].border = border

        statResult[0] = wrRowIdx - 2
        for r in range(2,wrRowIdx):
            readResult = ws.cell(r,4).value
            #print(readResult)
            if("Pass" in readResult):
                statResult[1] = statResult[1] +1
            elif("Fail" in readResult):
                statResult[2] = statResult[2] +1
            elif("Incon" in readResult):
                statResult[3] = statResult[3] +1
            else:
                statResult[4] = statResult[4] +1
        
        #print(i,statResult)
        workbookOutput.active = 0
        ws = workbookOutput.active 
        ws['C'+str(i+13)] = statResult[0]
        ws['D'+str(i+13)] = statResult[1]
        ws['E'+str(i+13)] = statResult[2]
        ws['F'+str(i+13)] = statResult[3]
        ws['G'+str(i+13)] = statResult[4]
        ws['I'+str(i+13)] = 100 - 100*statResult[4]/statResult[0]
        ws['J'+str(i+13)] = 100*statResult[1]/statResult[0]
    
    #sheet['A1'] = 'Hello Python, Hello Excel.'
    workbookOutput.save(finalReportName)
    #wb = openpyxl.load_workbook(finalReportName)
    #wb.save('new.xlsx')  
        
if __name__ == '__main__':
    AppStart()