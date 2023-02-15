## https://steam.oxxostudio.tw/category/python/example/excel-read.html 
## https://ithelp.ithome.com.tw/articles/10246377
## https://hackmd.io/@amostsai/SJkC1_EcX?type=view
## https://www.jb51.net/article/237279.htm
import os
import math
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border,Side
import csv
#from openpyxl.chart import LineChart, Reference
##import matplotlib.pyplot as plt
#from math import ceil
##===========================#    
#def fill_data_into_excel(r, c, data,sheet):
#    sheet.cell(r, c).value = data
#  
#def Init_Excel_Table(row, sheet):
#    colInfo = {}
#    for i in range(0,len(row)):
#        if(row[i] == "start_time"):
#            colInfo["start_time"] = (i)
#            sheet.cell(1, COL_INDEX_TIME).value = 'start_time'
#        elif(row[i] == "channel"):
#            colInfo["channel"] = (i)
#        #    sheet.write(0, COL_INDEX_CHANL_L, 'channel')
#        elif(row[i] == "data"):
#            colInfo["data"] = (i) 
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_HEX_L, 'HEX_24_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_HEX_R, 'HEX_24_R', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_DEC_L, 'DEC_24_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_DEC_R, 'DEC_24_R', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_16bit_DEC_L, 'DEC_16_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_16bit_DEC_R, 'DEC_16_R', sheet) 
#            #sheet.cell(1, COL_INDEX_DATA_24bit_HEX_L).value = 'HEX_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_HEX_R).value = 'HEX_24_R'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_DEC_L).value = 'DEC_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_DEC_R).value = 'DEC_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_16bit_DEC_L).value = 'DEC_16_L'
#            #sheet.cell(1, COL_INDEX_DATA_16bit_DEC_R).value = 'DEC_16_L'    
#    
#    fill_data_into_excel(1, COL_INDEX_SAMPLING_RATE, 'SamplingRate(Hz)', sheet)
#    #fill_data_into_excel(4, COL_INDEX_WAVEFORM, 'Waveform(L)', sheet)
#    
#    return colInfo;
#
#def convert_hexstr2hexval(valStr):
#    CONST_SIGN_BIT      = 0x800000
#    CONST_SIGN_VAL_MAX  = 0x7FFFFF
#    
#    try:
#        val_24 = int(valStr, 16)
#    except:
#        print(valStr)
#    if(val_24 & CONST_SIGN_BIT):
#        val_24 = (val_24 & CONST_SIGN_VAL_MAX) - CONST_SIGN_VAL_MAX - 1 
#    return val_24
#    
#
#def raw_convert_bytearr(rawList):
#    retList = []
#    for i in range(0, len(rawList)):
#        retList.append( rawList[i] & 0xFF)
#        retList.append( (rawList[i] >> 8) & 0xFF ) 
#    #print(retList)
#    return bytearray(retList)
#    
#def dual_raw_convert_bytearr(rawList1, rawList2):
#    retList = []
#    for i in range(0, len(rawList1)):
#        retList.append( rawList1[i] & 0xFF)
#        retList.append( (rawList1[i] >> 8) & 0xFF ) 
#        retList.append( rawList2[i] & 0xFF)
#        retList.append( (rawList2[i] >> 8) & 0xFF )        
#    #print(retList)
#    return bytearray(retList)
#
#def data_plot(ws, dataLen):
#    chart = LineChart()
#    chart.title = "Data - left, 24bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_24bit_DEC_L, min_row=1, max_col=COL_INDEX_DATA_24bit_DEC_L, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)
#    ws.add_chart(chart, "I4")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - right, 24bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_24bit_DEC_R, min_row=1, max_col=COL_INDEX_DATA_24bit_DEC_R, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)    
#    ws.add_chart(chart, "I20")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - left, 16bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_16bit_DEC_L, min_row=1, max_col=COL_INDEX_DATA_16bit_DEC_L, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)
#    ws.add_chart(chart, "I36")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - right, 16bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_16bit_DEC_R, min_row=1, max_col=COL_INDEX_DATA_16bit_DEC_R, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)    
#    ws.add_chart(chart, "I52")   
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10    
#
        
def AppStart():
    inConfigF = open("sourceFile.config", 'r')
    inConfigList = inConfigF.readlines()
    
#-----------------------------------------------------#    
    originalFile = inConfigList[0].split()[1]
    
    sheetInfo = [[inConfigList[3].split()[0], 0],
                 [inConfigList[4].split()[0], 0],
                 [inConfigList[5].split()[0], 0],
                 [inConfigList[6].split()[0], 0],
                 [inConfigList[7].split()[0], 0]
                 ]
    
    wb_original = openpyxl.load_workbook(originalFile)
    ws_original = wb_original["TestPlan"]
#-----------------------------------------------------#
    readRowIdx = 1
    for i in range(0,5):
        count = 0
        startFlag = False
        while(True):
            readResult = ws_original.cell(readRowIdx,1).value 
            if(readResult != None):
                if(startFlag == False):
                    if(sheetInfo[i][0] in readResult):
                        startFlag = True
                        readRowIdx = readRowIdx+2
                    else:
                        readRowIdx = readRowIdx+1
                else:
                    count = count+1
                    readRowIdx = readRowIdx+1
            else:
                if(startFlag == True):
                    print(sheetInfo[i][0],"\t", count)
                    readRowIdx = readRowIdx+1
                    break
                else:
                    readRowIdx = readRowIdx+1
                    
        
if __name__ == '__main__':
    AppStart()