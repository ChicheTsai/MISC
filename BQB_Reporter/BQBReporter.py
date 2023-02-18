## https://steam.oxxostudio.tw/category/python/example/excel-read.html 
## https://ithelp.ithome.com.tw/articles/10246377
## https://hackmd.io/@amostsai/SJkC1_EcX?type=view
## https://www.jb51.net/article/237279.htm
import os
import math
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border,Side
import csv
#from openpyxl.chart import LineChart, Reference
##import matplotlib.pyplot as plt
#from math import ceil
##===========================#    
#def fill_data_into_excel(r, c, data,sheet):
#    sheet.cell(r, c).value = data
#  
#def Init_Excel_Table(row, sheet):
#    colInfo = {}
#    for i in range(0,len(row)):
#        if(row[i] == "start_time"):
#            colInfo["start_time"] = (i)
#            sheet.cell(1, COL_INDEX_TIME).value = 'start_time'
#        elif(row[i] == "channel"):
#            colInfo["channel"] = (i)
#        #    sheet.write(0, COL_INDEX_CHANL_L, 'channel')
#        elif(row[i] == "data"):
#            colInfo["data"] = (i) 
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_HEX_L, 'HEX_24_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_HEX_R, 'HEX_24_R', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_DEC_L, 'DEC_24_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_24bit_DEC_R, 'DEC_24_R', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_16bit_DEC_L, 'DEC_16_L', sheet)
#            fill_data_into_excel(1, COL_INDEX_DATA_16bit_DEC_R, 'DEC_16_R', sheet) 
#            #sheet.cell(1, COL_INDEX_DATA_24bit_HEX_L).value = 'HEX_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_HEX_R).value = 'HEX_24_R'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_DEC_L).value = 'DEC_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_24bit_DEC_R).value = 'DEC_24_L'
#            #sheet.cell(1, COL_INDEX_DATA_16bit_DEC_L).value = 'DEC_16_L'
#            #sheet.cell(1, COL_INDEX_DATA_16bit_DEC_R).value = 'DEC_16_L'    
#    
#    fill_data_into_excel(1, COL_INDEX_SAMPLING_RATE, 'SamplingRate(Hz)', sheet)
#    #fill_data_into_excel(4, COL_INDEX_WAVEFORM, 'Waveform(L)', sheet)
#    
#    return colInfo;
#
#def convert_hexstr2hexval(valStr):
#    CONST_SIGN_BIT      = 0x800000
#    CONST_SIGN_VAL_MAX  = 0x7FFFFF
#    
#    try:
#        val_24 = int(valStr, 16)
#    except:
#        print(valStr)
#    if(val_24 & CONST_SIGN_BIT):
#        val_24 = (val_24 & CONST_SIGN_VAL_MAX) - CONST_SIGN_VAL_MAX - 1 
#    return val_24
#    
#
#def raw_convert_bytearr(rawList):
#    retList = []
#    for i in range(0, len(rawList)):
#        retList.append( rawList[i] & 0xFF)
#        retList.append( (rawList[i] >> 8) & 0xFF ) 
#    #print(retList)
#    return bytearray(retList)
#    
#def dual_raw_convert_bytearr(rawList1, rawList2):
#    retList = []
#    for i in range(0, len(rawList1)):
#        retList.append( rawList1[i] & 0xFF)
#        retList.append( (rawList1[i] >> 8) & 0xFF ) 
#        retList.append( rawList2[i] & 0xFF)
#        retList.append( (rawList2[i] >> 8) & 0xFF )        
#    #print(retList)
#    return bytearray(retList)
#
#def data_plot(ws, dataLen):
#    chart = LineChart()
#    chart.title = "Data - left, 24bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_24bit_DEC_L, min_row=1, max_col=COL_INDEX_DATA_24bit_DEC_L, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)
#    ws.add_chart(chart, "I4")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - right, 24bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_24bit_DEC_R, min_row=1, max_col=COL_INDEX_DATA_24bit_DEC_R, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)    
#    ws.add_chart(chart, "I20")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - left, 16bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_16bit_DEC_L, min_row=1, max_col=COL_INDEX_DATA_16bit_DEC_L, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)
#    ws.add_chart(chart, "I36")
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10
#
#    chart = LineChart()
#    chart.title = "Data - right, 16bit"
#    chart.style = 11
#    chart.height = 8
#    chart.width = 32    
#    chart.y_axis.title = 'value'
#    chart.x_axis.title = 'time'
#    data = Reference(ws, min_col=COL_INDEX_DATA_16bit_DEC_R, min_row=1, max_col=COL_INDEX_DATA_16bit_DEC_R, max_row = dataLen)
#    chart.add_data(data, titles_from_data=True)    
#    ws.add_chart(chart, "I52")   
#    s = chart.series[0]
#    s.smooth = True
#    s.graphicalProperties.line.width = 10    
#

def csv_to_excel(csv_filename, excel_filename):

    # Read CSV file
    csv_data = []
    with open(csv_filename) as f:
        csv_data = [row for row in csv.reader(f)]
    
    # Write to Excel file
    workbook = openpyxl.workbook.Workbook()
    ws = workbook.active
    
    lineCount = 1 
    numCol = 0
    sepSymbol = ""
    writeRow = 1
    keepReading = False
    readStr = ""
    for row in csv_data:
        if(lineCount == 1):
            sepSymbol = row[0][4]   #Assume the string begins with "sep=" 
        elif(lineCount == 2):
            numCol = row[0].count(';') +1
            #print(row[0],"\n")
        else:
            if(keepReading == False):
                readStr = ""
            
            for i in range(0,len(row)):
                readStr = readStr + row[i]
            readStr = readStr.replace("; ",". ")

            #if(writeRow >345 and writeRow <365):
            #    print(writeRow, keepReading, readStr.count(';'))
            #    print(readStr, "\n")
                
            if((keepReading == False) and (readStr.count(';') == 0)):
                keepReading = False
            elif((keepReading == False) and (readStr.count(';') +1 < numCol) ):
                keepReading = True
            elif((keepReading == True) and (readStr.count(';') +1 < numCol) ):
                keepReading = True
            else:
                keepReading = False
                readStr = readStr.split(";")
                #print(readStr)
                for c in range(0,len(readStr)):
                    ws[chr(65+c)+str(writeRow)].value =  readStr[c]
                writeRow = writeRow+1
            #print(lineCount, readStr)

            
        lineCount = lineCount+1
       
        #print(row)
        
        
        
    workbook.save(excel_filename)        
        
def find_info_from_report(itemName,wb_harmony, wb_ebq):
    ws_harmony = wb_harmony["TestPlan"]
    readRowIdx = 1
    while(True):
        readName = ws_harmony.cell(readRowIdx,1).value
        if(readName == itemName):
            if(ws_harmony.cell(readRowIdx,10).value == "Pass"):
                return ws_harmony.cell(readRowIdx,10).value
            else:
                break;
        
        readRowIdx += 1
        if(readRowIdx > 5000):
            break

    ws_ebq = wb_ebq["Sheet"]
    readRowIdx = 1
    while(True):
        readName = ws_ebq.cell(readRowIdx,1).value
        if(itemName == None):
            print("itemName")
        if(readName == None):
            return ""
        elif(itemName in readName):
            if(ws_ebq.cell(readRowIdx,5).value == "Passed"):
                return "Pass_EBQ"
            else:
                return ws_ebq.cell(readRowIdx,5).value
        
        readRowIdx += 1
        if(readRowIdx > 5000):
            break
            
#    for row in csvInRowEbq:
#        print(row)
#        print("\n\n")
    
#    while(True):
#        readName = ws_harmony.cell(readRowIdx,1).value
#        if(readName == itemName):
#            if(ws_harmony.cell(readRowIdx,10).value == "Pass"):
#                return ws_harmony.cell(readRowIdx,10).value
#            else:
#                break;
#        
#        readRowIdx += 1
#        if(readRowIdx > 5000):
#            break    
def BQB_itemCounting(fileName):
    wb_original = openpyxl.load_workbook(fileName)    
    ws_original = wb_original["TestPlan"]    
    
    readRowIdx = 13 #Skip the first 12 rows
    startCounting = False
    numOfCount = 0
    emptyCount = 0
    result = []
    tmpResult = ["",0]
    while(True):
        readData = ws_original.cell(readRowIdx,1).value
        #print(readRowIdx, startCounting == False, readData==None, emptyCount )
        
        if(startCounting == False):
            if(readData == None):
                emptyCount = emptyCount+1
            elif(readData == "Test Case ID"):
                startCounting = True
                emptyCount = 0
                tmpResult[0] = ws_original.cell(readRowIdx -1,1).value
        else:
            if(readData == None):
                tmpResult[1] = numOfCount
                #print(numOfCount, tmpResult[1])
                result.append(tmpResult)
                
                startCounting = False
                numOfCount = 0
                tmpResult = ["",0]
                
                
            else:
                numOfCount = numOfCount+1

        if(emptyCount > 10):
            for i in range(0, len(result)):
                print(result[i][0],"\t",result[i][1])
            return
        readRowIdx = readRowIdx+1

def isItemInOtherFile(itemName, ws):
    readRowIdx = 1     #Skip the first 12 rows
    emptyCount = 0
    while(True):
        readItem = ws.cell(readRowIdx,1).value
        if(readItem == None):
            emptyCount = emptyCount+1
        else:
            if(itemName in readItem):
                return True
            emptyCount = 0
        
        if(emptyCount > 10):
            return False
        readRowIdx = readRowIdx + 1
        
def BQB_versionComparator(fileName0,fileName1, tcrlVer):
    finalReportName = "CompareResult.xlsx"
    wb0 = openpyxl.load_workbook(fileName0)    
    ws0 = wb0["TestPlan"] 
    wb1 = openpyxl.load_workbook(fileName1)    
    ws1 = wb1["TestPlan"]
    
    workbookOutput = openpyxl.Workbook()
    ss_sheet = workbookOutput['Sheet']
    workbookOutput.active = 0
    ws = workbookOutput.active
    
    ws['A1'].value =  "Item"
    ws['B1'].value =  "New Item"
    ws['C1'].value =  "Removed Item"
    ws['D1'].value =  "Changed Item"
    ws.column_dimensions['A'].width = 25.0
    ws.column_dimensions['B'].width = 10.0
    ws.column_dimensions['C'].width = 14.0
    ws.column_dimensions['D'].width = 14.0
    
    
    # Check the newer item & changed items in newer test plan 
    writeRow = 2
    readRowIdx = 13     #Skip the first 12 rows
    startReading = False
    emptyCount = 0
    while(True):
        readItem = ws0.cell(readRowIdx,1).value
        #print(startReading == False, readItem == None,readRowIdx, emptyCount)
    
        if(startReading == False):
            if(readItem == None):
                emptyCount = emptyCount+1
            elif(readItem == "Test Case ID"):
                emptyCount = 0
                startReading = True
        else:
            if(readItem == None):
                startReading = False
            else:
                colBitfield = 0
                
                if(isItemInOtherFile(readItem,ws1) == False):
                    colBitfield = colBitfield | 0x01
                if( ws0.cell(readRowIdx,9).value != None):
                    if( tcrlVer in ws0.cell(readRowIdx,9).value):
                        colBitfield = colBitfield | 0x04
                
                if(colBitfield != 0):
                    ws["A"+str(writeRow)] = readItem
                    if(colBitfield & 0x01):
                        ws["B"+str(writeRow)] = "V"
                    if(colBitfield & 0x04):
                        ws["D"+str(writeRow)] = ws0.cell(readRowIdx,9).value
                    writeRow = writeRow+1
                    
        if(emptyCount > 10):
            break;
        readRowIdx = readRowIdx+1    
    
    # Check the removed item in newer test plan 
    readRowIdx = 13     #Skip the first 12 rows
    startReading = False
    emptyCount = 0
    while(True):
        readItem = ws1.cell(readRowIdx,1).value
        #print(startReading == False, readItem == None,readRowIdx, emptyCount)
    
        if(startReading == False):
            if(readItem == None):
                emptyCount = emptyCount+1
            elif(readItem == "Test Case ID"):
                emptyCount = 0
                startReading = True
        else:
            if(readItem == None):
                startReading = False
            else:
                colBitfield = 0
                
                if(isItemInOtherFile(readItem,ws0) == False):
                    colBitfield = colBitfield | 0x02
                
                if(colBitfield == 0x02):
                    ws["A"+str(writeRow)] = readItem
                    ws["C"+str(writeRow)] = "V"
                    writeRow = writeRow+1
                    
        if(emptyCount > 10):
            break;
        readRowIdx = readRowIdx+1    
    
    
    workbookOutput.save(finalReportName)
    
    
        
        
def BQB_ReportCombiner(fileName0,fileName1,fileName2):
    wb_original = openpyxl.load_workbook(fileName0)    
    ws_original = wb_original["TestPlan"]   

    wb_harmony = openpyxl.load_workbook(fileName1)
    
    # xlsx file,converted from CSV
    csvfile_ebq = open(fileName2)
    csv_to_excel(fileName2, "Tmp.xlsx")
    wb_ebq = openpyxl.load_workbook("Tmp.xlsx")
    
    sheetInfo = []
    # Part1: Finding How many categories   
    startCounting = False
    emptyCount = 0
    readRowIdx = 13     #Skip the first 12 rows
    while(True):
        readData = ws_original.cell(readRowIdx,1).value
        
        if(readData == None):
            emptyCount = emptyCount+1
        else:
            if(readData == "Test Case ID"):
                sheetInfo.append([ws_original.cell(readRowIdx-1,1).value, readRowIdx+1])
            emptyCount = 0
        
        if(emptyCount >10):
            #print(sheetInfo)
            break;
        readRowIdx = readRowIdx+1
    
    # Part2: Initialize the final report
    finalReportName = "FinalReport.xlsx"
    
    workbookOutput = openpyxl.Workbook()
    ss_sheet = workbookOutput['Sheet']
    ss_sheet.title = 'Summary'
    for i in range(0, len(sheetInfo)):
        workbookOutput.create_sheet(str(i+1)+"_"+sheetInfo[i][0], i+2)
    
    workbookOutput.active = 0
    ws = workbookOutput.active
    ws['B2'].value =  "Report Type"
    ws['B3'].value =  "Testing Purpose"
    ws['B4'].value =  "Release version"
    ws['B5'].value =  "Interface"
    ws['B6'].value =  "Host Platform"
    ws['B7'].value =  "Host OS"
    ws['B8'].value =  "Report location"
    
    ws['B10'].value =  "Suites Breakdown"
    ws['C10'].value =  "Total"
    ws['D10'].value =  "Pass"
    ws['E10'].value =  "Fail"
    ws['F10'].value =  "Inconclusive"
    ws['G10'].value =  "NA"
    ws['H10'].value =  ""
    ws['I10'].value =  "Completed[%]"
    ws['J10'].value =  "Pass[%]"
    ws['K10'].value =  "Bug Count"

    for i in range(0, len(sheetInfo)):
        ws['B'+str(i+11)].value = sheetInfo[i][0]
    
    jiraStartRow = 10+2+len(sheetInfo)
    ws['B'+str(jiraStartRow)].value = "JIRA" 
    ws['C'+str(jiraStartRow)].value = "Summary"						
    ws['I'+str(jiraStartRow)].value = "Priority"
    ws['J'+str(jiraStartRow)].value = "Status"
    ws['K'+str(jiraStartRow)].value = "Assignee"
    
    ws.column_dimensions['B'].width = 16.5
    ws.column_dimensions['C'].width = 10.0
    ws.column_dimensions['D'].width = 10.0
    ws.column_dimensions['E'].width = 10.0
    ws.column_dimensions['F'].width = 10.0
    ws.column_dimensions['G'].width = 8.0
    ws.column_dimensions['I'].width = 14.0
    ws.column_dimensions['J'].width = 14.0
    ws.column_dimensions['K'].width = 14.0
    ws.row_dimensions[4].height = 40.0
    
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=11)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=11)
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=11)
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=11)
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=11)
    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=11)
    ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=11)
    ws.merge_cells(start_row=jiraStartRow+0, start_column=3, end_row=jiraStartRow+0, end_column=8)
    ws.merge_cells(start_row=jiraStartRow+1, start_column=3, end_row=jiraStartRow+1, end_column=8)
    ws.merge_cells(start_row=jiraStartRow+2, start_column=3, end_row=jiraStartRow+2, end_column=8)
    ws.merge_cells(start_row=jiraStartRow+3, start_column=3, end_row=jiraStartRow+3, end_column=8)
    ws.merge_cells(start_row=jiraStartRow+4, start_column=3, end_row=jiraStartRow+4, end_column=8)
    ws.merge_cells(start_row=jiraStartRow+5, start_column=3, end_row=jiraStartRow+5, end_column=8)
    
    border = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000'))
    
    for c in range(2,12):                    
        for r in range(2,9):
            ws[chr(65+c-1)+str(r)].border = border
        for r in range(10,10+len(sheetInfo)+1):
            ws[chr(65+c-1)+str(r)].border = border
        for r in range(jiraStartRow,jiraStartRow+6):
            ws[chr(65+c-1)+str(r)].border = border
    
    for i in range(1,len(sheetInfo)+1):
        workbookOutput.active = i
        ws = workbookOutput.active    
        ws['A1'].value =  "Item"
        ws['B1'].value =  "Description"
        ws['C1'].value =  "TestPlatform"
        ws['D1'].value =  "Result"
        ws['E1'].value =  "JIRA"        #Filled by hand if needed, QA 
        ws['F1'].value =  "Reason"      #Filled by hand if needed, FW
        ws['G1'].value =  "Pass rate"   #Filled by hand
        ws['H1'].value =  "New Item"    #Filled by hand if needed, QA
        ws.column_dimensions['A'].width = 25.0
        ws.column_dimensions['B'].width = 30.0
        ws.column_dimensions['C'].width = 12.0
        ws.column_dimensions['D'].width = 12.0
        ws.column_dimensions['E'].width = 12.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 8.0
        ws.column_dimensions['H'].width = 9.0

    
    fontFail = Font(color='9c0006')
    fontInc = Font(color='006100')
    for i in range(0,len(sheetInfo)):
        #print(i)
        statResult = [0,0,0,0,0]    #Total / Pass / Fail / Inconclusive / NA
        readRowIdx = sheetInfo[i][1]
        
        workbookOutput.active = i + 1
        ws = workbookOutput.active 
        wrRowIdx = 2

        # Part3: Fill the result & test platform 
        while(True):
            #print(ws_original.cell(readRowIdx,1).value)
            #Item
            ws['A'+str(wrRowIdx)] = ws_original.cell(readRowIdx,1).value
            #Description
            ws['B'+str(wrRowIdx)] = ws_original.cell(readRowIdx,2).value
            #TestPlatform
            ws['C'+str(wrRowIdx)] = ws_original.cell(readRowIdx,5).value
            #Result
            ws['D'+str(wrRowIdx)] = find_info_from_report(ws_original.cell(readRowIdx,1).value, wb_harmony, wb_ebq)
            if("Fail" in ws['D'+str(wrRowIdx)].value):
                ws['D'+str(wrRowIdx)].font = fontFail
            elif("Inconclusive" in ws['D'+str(wrRowIdx)].value):
                ws['D'+str(wrRowIdx)].font = fontInc
                
            readRowIdx += 1
            wrRowIdx += 1

            if(ws_original.cell(readRowIdx,1).value == None):
                #print(readRowIdx, ws_original.cell(readRowIdx,1).value)
                break;
                
        # Part4: Configure the boarder  
        for c in range(1,9):
            for r in range(1,wrRowIdx):
                ws[chr(65+c-1)+str(r)].border = border

        # Part5: Calculate the result information  
        statResult[0] = wrRowIdx - 2
        for r in range(2,wrRowIdx):
            readResult = ws.cell(r,4).value
            #print(readResult)
            if("Pass" in readResult):
                statResult[1] = statResult[1] +1
            elif("Fail" in readResult):
                statResult[2] = statResult[2] +1
            elif("Incon" in readResult):
                statResult[3] = statResult[3] +1
            else:
                statResult[4] = statResult[4] +1
        
        #print(i,statResult)
        # Part6: Calculate the stac-result
        workbookOutput.active = 0
        ws = workbookOutput.active 
        ws['C'+str(i+11)] = statResult[0]
        ws['D'+str(i+11)] = statResult[1]
        ws['E'+str(i+11)] = statResult[2]
        ws['F'+str(i+11)] = statResult[3]
        ws['G'+str(i+11)] = statResult[4]
        ws['I'+str(i+11)] = 100 - 100*statResult[4]/statResult[0]
        ws['J'+str(i+11)] = 100*statResult[1]/statResult[0]
        
        
    workbookOutput.save(finalReportName)

        
def AppStart():
#-----------------------------------------------------#    
    modeSelection = int(sys.argv[1])
    if(modeSelection == 0): #Itemcounting
        fileName = (sys.argv[2])
        BQB_itemCounting(fileName)
    elif(modeSelection == 1): #Item comparison
        fileName0 = (sys.argv[2])   #Newer test plan from BT SIG, xlsx
        fileName1 = (sys.argv[3])   #previous test plan from BT SIG, xlsx  
        tcrlVer = (sys.argv[4])
        BQB_versionComparator(fileName0, fileName1, tcrlVer)
    elif(modeSelection == 2):
        fileName0 = (sys.argv[2])   #Test plan from BT SIG, xlsx
        fileName1 = (sys.argv[3])   #Report Harmony, xlsx
        fileName2 = (sys.argv[4])   #Report EBQ, cvs
        BQB_ReportCombiner(fileName0,fileName1,fileName2)
    else:
        print("The mode is not supported now")
        exit()

        
if __name__ == '__main__':
    AppStart()